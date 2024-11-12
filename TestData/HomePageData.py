import openpyxl
from bokeh.layouts import column



class HomePageData:
    test_homepage_data = [{"firstname":"ankit", "email":"ankit@gmail.com"},{"firstname":"ekka", "email":"ekka@gmail.com"},{"firstname":"rahul", "email":"rahul@gmail.com"}]


    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("D:\Selenium\SeleniumSelProject\TestData\excel_test.xlsx")

        sheet = book.active

        cell = sheet.cell(row=1, column=2)

        # print(cell.value)
        Dict = {}

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]